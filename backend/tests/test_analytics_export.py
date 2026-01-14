"""
Tests for Analytics Export API endpoint.

TDD RED Phase: These tests should FAIL until implementation is complete.
"""
import pytest
import csv
import io
from datetime import date
from decimal import Decimal
from openpyxl import load_workbook
from app.models import CashDifference


class TestAnalyticsExportEndpoint:
    """Test the analytics export API endpoint."""

    def test_export_endpoint_exists(self, client):
        """
        Export endpoint should exist and be accessible.
        """
        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={
                "start_date": "2025-01-01",
                "end_date": "2025-01-31",
                "format": "csv"
            },
            headers={"X-Branch-Id": "1"}
        )

        # Should not be 404
        assert response.status_code != 404, "Export endpoint must exist"
        # Should be 200 or at least a valid response
        assert response.status_code == 200

    def test_export_requires_date_range(self, client):
        """
        Export endpoint should return 422 if date range is missing.
        """
        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={"format": "csv"},
            headers={"X-Branch-Id": "1"}
        )

        assert response.status_code == 422

    def test_export_requires_format_parameter(self, client):
        """
        Export endpoint should require format parameter.
        """
        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={
                "start_date": "2025-01-01",
                "end_date": "2025-01-31"
            },
            headers={"X-Branch-Id": "1"}
        )

        # Should fail without format - either 422 or default to csv (200)
        assert response.status_code in [200, 422]


class TestCsvExport:
    """Test CSV export functionality."""

    def test_csv_export_returns_csv_content_type(self, client, db):
        """
        CSV export should return correct content type.
        """
        # Create test data
        cd = CashDifference(
            branch_id=1,
            difference_date=date(2025, 1, 15),
            kasa_total=Decimal("1500"),
            pos_total=Decimal("1530"),
            status="pending",
            created_by=1
        )
        db.add(cd)
        db.commit()

        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={
                "start_date": "2025-01-01",
                "end_date": "2025-01-31",
                "format": "csv"
            },
            headers={"X-Branch-Id": "1"}
        )

        assert response.status_code == 200
        assert "text/csv" in response.headers.get("content-type", "")

    def test_csv_export_has_content_disposition(self, client, db):
        """
        CSV export should have Content-Disposition header for download.
        """
        cd = CashDifference(
            branch_id=1,
            difference_date=date(2025, 1, 15),
            kasa_total=Decimal("1500"),
            pos_total=Decimal("1530"),
            status="pending",
            created_by=1
        )
        db.add(cd)
        db.commit()

        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={
                "start_date": "2025-01-01",
                "end_date": "2025-01-31",
                "format": "csv"
            },
            headers={"X-Branch-Id": "1"}
        )

        assert response.status_code == 200
        content_disp = response.headers.get("content-disposition", "")
        assert "attachment" in content_disp
        assert ".csv" in content_disp

    def test_csv_export_contains_data(self, client, db):
        """
        CSV export should contain the analytics data.
        """
        cd = CashDifference(
            branch_id=1,
            difference_date=date(2025, 1, 15),
            kasa_total=Decimal("1500"),
            pos_total=Decimal("1530"),
            status="pending",
            created_by=1
        )
        db.add(cd)
        db.commit()

        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={
                "start_date": "2025-01-01",
                "end_date": "2025-01-31",
                "format": "csv"
            },
            headers={"X-Branch-Id": "1"}
        )

        assert response.status_code == 200

        # Parse CSV
        csv_content = response.content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(csv_content))
        rows = list(reader)

        # Should have at least one row
        assert len(rows) >= 1

        # Should contain our date
        dates = [row.get("date") or row.get("Tarih") for row in rows]
        assert "2025-01-15" in dates

    def test_csv_export_respects_tenant_isolation(self, client, db):
        """
        CSV export must respect tenant (branch) isolation.
        """
        # Create data for branch 1
        cd1 = CashDifference(
            branch_id=1,
            difference_date=date(2025, 1, 10),
            kasa_total=Decimal("1000"),
            pos_total=Decimal("1000"),
            status="pending",
            created_by=1
        )
        # Create data for branch 2 (should NOT appear in export)
        cd2 = CashDifference(
            branch_id=2,
            difference_date=date(2025, 1, 11),
            kasa_total=Decimal("9999"),
            pos_total=Decimal("9999"),
            status="pending",
            created_by=1
        )
        db.add_all([cd1, cd2])
        db.commit()

        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={
                "start_date": "2025-01-01",
                "end_date": "2025-01-31",
                "format": "csv"
            },
            headers={"X-Branch-Id": "1"}
        )

        assert response.status_code == 200

        csv_content = response.content.decode("utf-8")

        # Should contain branch 1 data
        assert "1000" in csv_content
        # Should NOT contain branch 2 data
        assert "9999" not in csv_content


class TestExcelExport:
    """Test Excel export functionality."""

    def test_excel_export_returns_xlsx_content_type(self, client, db):
        """
        Excel export should return correct content type.
        """
        cd = CashDifference(
            branch_id=1,
            difference_date=date(2025, 1, 15),
            kasa_total=Decimal("1500"),
            pos_total=Decimal("1530"),
            status="pending",
            created_by=1
        )
        db.add(cd)
        db.commit()

        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={
                "start_date": "2025-01-01",
                "end_date": "2025-01-31",
                "format": "excel"
            },
            headers={"X-Branch-Id": "1"}
        )

        assert response.status_code == 200
        content_type = response.headers.get("content-type", "")
        assert "spreadsheet" in content_type or "excel" in content_type or "octet-stream" in content_type

    def test_excel_export_has_content_disposition(self, client, db):
        """
        Excel export should have Content-Disposition header for download.
        """
        cd = CashDifference(
            branch_id=1,
            difference_date=date(2025, 1, 15),
            kasa_total=Decimal("1500"),
            pos_total=Decimal("1530"),
            status="pending",
            created_by=1
        )
        db.add(cd)
        db.commit()

        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={
                "start_date": "2025-01-01",
                "end_date": "2025-01-31",
                "format": "excel"
            },
            headers={"X-Branch-Id": "1"}
        )

        assert response.status_code == 200
        content_disp = response.headers.get("content-disposition", "")
        assert "attachment" in content_disp
        assert ".xlsx" in content_disp

    def test_excel_export_is_valid_xlsx(self, client, db):
        """
        Excel export should produce a valid XLSX file.
        """
        cd = CashDifference(
            branch_id=1,
            difference_date=date(2025, 1, 15),
            kasa_total=Decimal("1500"),
            pos_total=Decimal("1530"),
            status="pending",
            created_by=1
        )
        db.add(cd)
        db.commit()

        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={
                "start_date": "2025-01-01",
                "end_date": "2025-01-31",
                "format": "excel"
            },
            headers={"X-Branch-Id": "1"}
        )

        assert response.status_code == 200

        # Should be loadable by openpyxl
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Should have data rows
        assert ws.max_row >= 2  # Header + at least 1 data row

    def test_excel_export_contains_data(self, client, db):
        """
        Excel export should contain the analytics data.
        """
        cd = CashDifference(
            branch_id=1,
            difference_date=date(2025, 1, 15),
            kasa_total=Decimal("1500"),
            pos_total=Decimal("1530"),
            status="pending",
            created_by=1
        )
        db.add(cd)
        db.commit()

        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={
                "start_date": "2025-01-01",
                "end_date": "2025-01-31",
                "format": "excel"
            },
            headers={"X-Branch-Id": "1"}
        )

        assert response.status_code == 200

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Find all cell values
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(v) for v in row if v is not None])

        # Should contain our date somewhere
        assert any("2025-01-15" in v for v in all_values)


class TestExportEdgeCases:
    """Test edge cases for export."""

    def test_export_empty_date_range(self, client):
        """
        Empty date range should return valid empty file, not error.
        """
        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={
                "start_date": "2099-01-01",
                "end_date": "2099-01-31",
                "format": "csv"
            },
            headers={"X-Branch-Id": "1"}
        )

        # Should succeed even with no data
        assert response.status_code == 200

    def test_invalid_format_returns_error(self, client):
        """
        Invalid format parameter should return 422.
        """
        response = client.get(
            "/api/reports/daily-sales-analytics/export",
            params={
                "start_date": "2025-01-01",
                "end_date": "2025-01-31",
                "format": "invalid_format"
            },
            headers={"X-Branch-Id": "1"}
        )

        assert response.status_code == 422
