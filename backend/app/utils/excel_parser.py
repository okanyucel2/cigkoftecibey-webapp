"""
Excel Kasa Raporu Parser

Parses the cashier's Excel report (1453.xlsx format) and extracts:
- Sales data by channel (VISA, NAKİT, online platforms)
- Expense items from the GIDERLER section
"""
from datetime import date, datetime
from decimal import Decimal
from typing import Optional
import openpyxl
from io import BytesIO


def parse_kasa_raporu(file_content: bytes) -> dict:
    """
    Parse Excel Kasa Raporu and return structured data.

    Expected Excel structure (FORM sheet):
    - Row 4: TARIH (Date)
    - Row 6: VISA
    - Row 7: PAKET VISA
    - Row 8: NFS KOMISYON
    - Row 10: NAKIT
    - Row 11: PAKET NAKIT
    - Row 13: TRENDYOL
    - Row 14: GETIR
    - Row 15: YEMEK SEPETI
    - Row 16: MIGROS YEMEK
    - Rows 21+: GIDERLER (expenses)

    Columns: D = Sabahci, N = Aksamci
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)

    # Try FORM sheet first, then KASA RAPORU
    if 'FORM' in wb.sheetnames:
        sheet = wb['FORM']
        return _parse_form_sheet(sheet)
    elif 'KASA RAPORU' in wb.sheetnames:
        sheet = wb['KASA RAPORU']
        return _parse_kasa_raporu_sheet(sheet)
    else:
        raise ValueError("Excel dosyasinda FORM veya KASA RAPORU sayfasi bulunamadi")


def _parse_form_sheet(sheet) -> dict:
    """Parse FORM sheet with two shifts (Sabahci + Aksamci)"""

    def get_value(row: int, col: int) -> Decimal:
        """Get cell value, handle formulas by getting cached value"""
        val = sheet.cell(row=row, column=col).value
        if val is None:
            return Decimal("0")
        if isinstance(val, str) and val.startswith('='):
            return Decimal("0")
        try:
            return Decimal(str(val))
        except:
            return Decimal("0")

    # Column indices: D=4 (Sabahci), N=14 (Aksamci)
    COL_SABAHCI = 4
    COL_AKSAMCI = 14

    # Parse date from row 4
    date_val = sheet.cell(row=4, column=COL_SABAHCI).value
    if isinstance(date_val, datetime):
        parsed_date = date_val.date()
    elif isinstance(date_val, date):
        parsed_date = date_val
    else:
        parsed_date = date.today()

    # Sum both shifts
    result = {
        "date": parsed_date,
        "visa": (
            get_value(6, COL_SABAHCI) + get_value(6, COL_AKSAMCI) +
            get_value(7, COL_SABAHCI) + get_value(7, COL_AKSAMCI) +
            get_value(8, COL_SABAHCI) + get_value(8, COL_AKSAMCI)
        ),
        "nakit": (
            get_value(10, COL_SABAHCI) + get_value(10, COL_AKSAMCI) +
            get_value(11, COL_SABAHCI) + get_value(11, COL_AKSAMCI)
        ),
        "trendyol": get_value(13, COL_SABAHCI) + get_value(13, COL_AKSAMCI),
        "getir": get_value(14, COL_SABAHCI) + get_value(14, COL_AKSAMCI),
        "yemeksepeti": get_value(15, COL_SABAHCI) + get_value(15, COL_AKSAMCI),
        "migros": get_value(16, COL_SABAHCI) + get_value(16, COL_AKSAMCI),
        "expenses": []
    }

    # Parse expenses (rows 21+) - IMPORTANT: These are deducted from NAKIT
    # So when comparing with POS, we need to add expenses back to NAKIT
    expenses_total = Decimal("0")
    for row in range(21, sheet.max_row + 1):
        desc = sheet.cell(row=row, column=1).value
        amount_s = get_value(row, 8)
        amount_a = get_value(row, 18)

        if desc and (amount_s > 0 or amount_a > 0):
            exp_amount = amount_s + amount_a
            expenses_total += exp_amount
            result["expenses"].append({
                "description": str(desc),
                "amount": exp_amount
            })

    # Add expenses to nakit for accurate comparison with POS
    # (In the KASA report, expenses reduce NAKIT, but POS shows the full amount)
    result["nakit"] = result["nakit"] + expenses_total

    result["total"] = (
        result["visa"] + result["nakit"] +
        result["trendyol"] + result["getir"] +
        result["yemeksepeti"] + result["migros"]
    )

    return result


def _parse_kasa_raporu_sheet(sheet) -> dict:
    """Parse KASA RAPORU sheet (combined summary)"""

    def get_value(row: int, col: int = 4) -> Decimal:
        val = sheet.cell(row=row, column=col).value
        if val is None:
            return Decimal("0")
        try:
            return Decimal(str(val))
        except:
            return Decimal("0")

    date_val = sheet.cell(row=4, column=4).value
    if isinstance(date_val, datetime):
        parsed_date = date_val.date()
    elif isinstance(date_val, date):
        parsed_date = date_val
    else:
        parsed_date = date.today()

    result = {
        "date": parsed_date,
        "visa": get_value(6) + get_value(7) + get_value(8),
        "nakit": get_value(10) + get_value(11),
        "trendyol": get_value(13),
        "getir": get_value(14),
        "yemeksepeti": get_value(15),
        "migros": get_value(16),
        "expenses": []
    }

    # Parse expenses (rows 21+) - IMPORTANT: Add to nakit for POS comparison
    expenses_total = Decimal("0")
    for row in range(21, sheet.max_row + 1):
        desc = sheet.cell(row=row, column=1).value
        amount = get_value(row, 8)

        if desc and amount > 0:
            expenses_total += amount
            result["expenses"].append({
                "description": str(desc),
                "amount": amount
            })

    # Add expenses to nakit for accurate comparison with POS
    result["nakit"] = result["nakit"] + expenses_total

    result["total"] = (
        result["visa"] + result["nakit"] +
        result["trendyol"] + result["getir"] +
        result["yemeksepeti"] + result["migros"]
    )

    return result


def parse_hasilat_raporu(file_content: bytes) -> dict:
    """
    Parse Şefim Hasılat Raporu Excel and return structured data.

    Expected Excel structure (values are in column D):
    - Row 3: TARİH (date value)
    - Row 4: VISA
    - Row 5: POS 1
    - Row 6: POS 2
    - Row 7: PAKET VISA
    - Row 8: FATURALI SATIŞ
    - Row 10: NAKİT
    - Row 11: PAKET NAKİT
    - Row 13: TRENDYOL
    - Row 14: GETİR
    - Row 15: YEMEK SEPETİ
    - Row 16: MİGROS YEMEK
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    sheet = wb.active  # Use first sheet

    def get_value(row: int, col: int = 4) -> Decimal:
        """Get cell value as Decimal. Default column is 4 (column D)."""
        val = sheet.cell(row=row, column=col).value
        if val is None:
            return Decimal("0")
        try:
            return Decimal(str(val))
        except:
            return Decimal("0")

    # Parse date from row 3 (column 4 = D column has the date value)
    date_val = sheet.cell(row=3, column=4).value
    if isinstance(date_val, datetime):
        parsed_date = date_val.date()
    elif isinstance(date_val, date):
        parsed_date = date_val
    elif isinstance(date_val, (int, float)):
        # Excel serial date format (e.g., 46014 = 2026-01-14)
        parsed_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_val) - 2).date()
    else:
        parsed_date = date.today()

    # VISA breakdown: VISA + POS 1 + POS 2 + PAKET VISA + FATURALI
    visa_total = (
        get_value(4) +  # VISA
        get_value(5) +  # POS 1
        get_value(6) +  # POS 2
        get_value(7) +  # PAKET VISA
        get_value(8)    # FATURALI SATIŞ
    )

    # NAKİT: NAKİT + PAKET NAKİT
    nakit_total = get_value(10) + get_value(11)

    result = {
        "date": parsed_date,
        "visa": visa_total,
        "nakit": nakit_total,
        "trendyol": get_value(13),
        "getir": get_value(14),
        "yemeksepeti": get_value(15),  # YEMEK SEPETİ is in the report
        "migros": get_value(16),
    }

    result["total"] = (
        result["visa"] + result["nakit"] +
        result["trendyol"] + result["getir"] +
        result["yemeksepeti"] + result["migros"]
    )

    return result
